import openpyxl 
from datetime import datetime
cells = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
def get_dict():
    path = "Attendance.xlsx"
    wb_obj = openpyxl.load_workbook(path) 
    sheet_obj = wb_obj.active 
    max_r = sheet_obj.max_row
    print("Line 9:",max_r)
    cell_obj = sheet_obj['A1': 'B'+str(max_r)]
    dict = {}
    for cell1, cell2 in cell_obj:
        dict[cell1.row] = [cell2.value, cell1.value]
    print(dict)
    return dict

def add_column():
    today = datetime.date(datetime.now())
    print(today)
    path = "Attendance.xlsx"
    wb_obj = openpyxl.load_workbook(path) 
    sheet_obj = wb_obj.active 
    last = sheet_obj.max_column
    last_column_name = get_col_name(last)
    if last_column_name == str(today):
        print("Already added")
    else:
        cell = sheet_obj.cell(row=1, column=last+1)
        cell.value = str(today)
        # initialise all values in the new column as Absent
        for i in range(2, sheet_obj.max_row+1):
            cell = sheet_obj.cell(row=i, column=last+1)
            cell.value = 'A'
        wb_obj.save(path)
    return last
        
def get_col_name(colnum):
    path = "Attendance.xlsx"
    wb_obj = openpyxl.load_workbook(path) 
    sheet_obj = wb_obj.active 
    cell = sheet_obj.cell(row=1, column=colnum)
    return cell.value

def write_attendance_in_xl(name, data):
    roll_no = int(name.split('-')[0])
    print("on line 44:",roll_no)
    stud_name = name.split('-')[1]
    lastColumn = add_column()
    path = "Attendance.xlsx"
    wb_obj = openpyxl.load_workbook(path) 
    sheet_obj = wb_obj.active 
    for key, value in data.items():
        if stud_name in value and roll_no in value:
            cell = sheet_obj.cell(row=key, column=lastColumn)
            print("on line 56:",cell.value, cell)
            if cell.value == "A":
                cell.value = 'P ' + str(datetime.now().time())
    wb_obj.save(path)