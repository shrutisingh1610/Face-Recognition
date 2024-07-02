import cv2
import os
import time
import openpyxl

def add_the_face():
    ######### if these directory doesn't exist, make them #######
    try:os.makedirs('images')
    except:pass

    ############### take name and roll no as input #################
    name = input('Enter your name --> ')
    roll = input('Enter your roll no --> ')
    while not roll.isdigit():
        roll = input('Enter your roll no --> ')

    imageName = roll+'-'+name

    ######### add image in folder ##################
    if imageName + '.png' in os.listdir('images') or imageName + '.jpg' in os.listdir('images'):
        print('User already exists')

    else:    
        cap = cv2.VideoCapture(0)
        i = 0
        print()
        for i in range(5):
            print(f'Capturing starts in {5-i} seconds...')
            time.sleep(1)
        print('Taking photo...')

        ret, frame = cap.read()
        cv2.imshow("Taking your picture, be steady", frame)
        if cv2.imwrite('images/' + imageName + '.png', frame):
            print('Image saved successfully')
            try:
                workbook = openpyxl.load_workbook('Attendance.xlsx')
                sheet = workbook.active
                last_row = sheet.max_row
                last_column = sheet.max_column
                sheet.cell(row=last_row+1, column=1).value = name
                sheet.cell(row=last_row+1, column=2).value = int(roll)
                for i in range(3, last_column+1):
                    sheet.cell(row=last_row+1, column=i).value = 'A'
                workbook.save('Attendance.xlsx')
            except:
                print('Error in saving data.\nMust be the following reasons: \n1. Attendance.xlsx is not present in the current directory\n2. Attendance.xlsx is not in the correct format\n3. Roll number is not a number')
        else:
            print('Image not saved\nPlease try again')
        cv2.destroyAllWindows()
        cap.release()